# test_creditscalc

import unittest
import creditscalc
from openpyxl import Workbook


class TestCreditsCalc(unittest.TestCase):

    def test_room(self):
        # rates = dict(Desk='1', Meeting='3', Gallery='5', Theatre='20', Outdoor='20')
        wb = self.create_test_wb()
        # ws = wb['Rates']
        # creditscalc.rates = creditscalc.load_rates(ws)
        ffh = creditscalc.Room('First Floor Hall', 'Gallery')
        self.assertEqual('First Floor Hall', ffh.name)
        self.assertEqual('Gallery', ffh.type)
        # self.assertEqual('5', ffh.get_credit())

    def test_get_rate(self):
        wb = self.create_test_wb()
        ws = wb['Rates']
        rates = creditscalc.load_rates(ws)
        ffh = creditscalc.Room('First Floor Hall', 'Gallery')
        self.assertEqual('5', creditscalc.get_rate(ffh.type, rates))

    def test_load_rooms(self):
        wb = self.create_test_wb()
        ws = wb['Rates']
        creditscalc.rates = creditscalc.load_rates(ws)
        ws = wb["Facilities"]
        rooms = creditscalc.load_rooms(ws)
        # self.assertEqual({'First Floor Hall': {'name': 'First Floor Hall','type':'Gallery'},
        #                   'Second Floor Hall': {'name': 'Second Floor Hall', 'type': 'Gallery'},
        #                   'Third Floor Hall': {'name': 'Third Floor Hall', 'type': 'Gallery'},
        #                   'RGO Treehouse (South)': {'name': 'RGO Treehouse (South)', 'type': 'Meeting'},
        #                   'RGO Treehouse (North)': {'name': 'RGO Treehouse (North)', 'type': 'Meeting'},
        #                   'RGO Treehouse (Full)': {'name': 'RGO Treehouse (Full)', 'type': 'Meeting'},
        #                   'Studio (North)': {'name': 'Studio (North)', 'type': 'Theatre'},
        #                   'Studio (South)': {'name': 'Studio (South)', 'type': 'Theatre'},
        #                   'ArtPark': {'name': 'ArtPark', 'type': 'Outdoor'},
        #                   'Desk 3rd floor 1': {'name': 'Desk 3rd floor 1', 'type': 'Desk'},
        #                   'Desk 3rd floor 2': {'name': 'Desk 3rd floor 2', 'type': 'Desk'},
        #                   'Desk 3rd floor 3': {'name': 'Desk 3rd floor 3', 'type': 'Desk'},
        #                   'Desk 3rd floor 4': {'name': 'Desk 3rd floor 4', 'type': 'Desk'},
        #                   'Desk 3rd floor 5': {'name': 'Desk 3rd floor 5', 'type': 'Desk'},
        #                   'Desk 3rd floor 6': {'name': 'Desk 3rd floor 6', 'type': 'Desk'},
        #                   'Desk 3rd floor 7': {'name': 'Desk 3rd floor 7', 'type': 'Desk'},
        #                   'Desk 3rd floor 8': {'name': 'Desk 3rd floor 8', 'type': 'Desk'},
        #                   'Desk 3rd floor 9': {'name': 'Desk 3rd floor 9', 'type': 'Desk'}}, rooms)
        self.assertEqual('Desk 3rd floor 2', rooms['Desk 3rd floor 2'].name)
        self.assertEqual('Desk', rooms['Desk 3rd floor 2'].type)
        # self.assertEqual('1', rooms['Desk 3rd floor 2'].get_credit())

    def test_client(self):
        clt1 = creditscalc.Client('Eric Day')
        self.assertEqual('Eric Day', clt1.name)
        self.assertEqual('Eric', clt1.first_name)
        self.assertEqual('Day', clt1.last_name)
        self.assertEqual(0, clt1.credits)
        clt1.add_credit(10)
        self.assertEqual(10, clt1.credits)
        clt1.add_credit(10)
        clt1.add_credit(10)
        clt1.add_credit(10)
        self.assertEqual(40, clt1.credits)

    def test_cell_not_empty(self):
        self.assertTrue(creditscalc.cell_not_empty('Carri'))
        self.assertFalse(creditscalc.cell_not_empty(''))
        self.assertFalse(creditscalc.cell_not_empty(None))

    def test_valid_client(self):
        wb = self.create_test_wb()
        ws = wb["Clients"]
        clients = creditscalc.load_clients(ws)
        self.assertTrue(creditscalc.valid_client(clients, 'Carri'))
        self.assertFalse(creditscalc.valid_client(clients, 'Eric'))

    def test_load_clients(self):
        wb = self.create_test_wb()
        ws = wb["Clients"]
        clients = creditscalc.load_clients(ws)
        # self.assertEqual({'Carri': {'name': 'Carri Cordon','first_name':'Carri', 'last_name': 'Cordon', 'credits': 0},
        #                   'Caren': {'name': 'Caren Alsop','first_name':'Caren', 'last_name': 'Alsop', 'credits': 0}}, clients)
        self.assertEqual('Carri Cordon', clients['Carri'].name)
        self.assertEqual(0, clients['Carri'].get_credits())
        clients['Carri'].add_credit(20)
        self.assertEqual(20, clients['Carri'].get_credits())

    def test_load_rates(self):
        wb = self.create_test_wb()
        ws = wb['Rates']
        mrates = creditscalc.load_rates(ws)
        self.assertEqual(mrates, {'Desk':'1', 'Meeting': '3', 'Gallery': '5', 'Theatre': '20', 'Outdoor': '20'})
        self.assertEqual('1', mrates['Desk'])

    def test_check_client_use(self):
        wb = self.create_test_wb()
        ws = wb['Rates']
        mrates = creditscalc.load_rates(ws)
        ws = wb["Facilities"]
        rooms = creditscalc.load_rooms(ws)
        ws = wb["Clients"]
        clients = creditscalc.load_clients(ws)
        result = creditscalc.check_client_use(wb, '2018-08', clients, mrates, rooms)
        self.assertIsNone(result)

    def test_print_clients(self):
        wb = self.create_test_wb()
        ws = wb['Rates']
        mrates = creditscalc.load_rates(ws)
        ws = wb["Facilities"]
        rooms = creditscalc.load_rooms(ws)
        ws = wb["Clients"]
        clients = creditscalc.load_clients(ws)
        result = creditscalc.check_client_use(wb, '2018-08', clients, mrates, rooms)
        creditscalc.print_clients(clients)


    def test_open_my_workbook(self):
        wb = self.create_test_wb()
        file_name = 'cSpace_Booking.xlsx'
        self.assertIsNotNone(wb, creditscalc.open_my_workbook(file_name))

    # def test_client(self):
    #     client1 = creditscalc.Client('Larry', 'Shumlich', 29)
    #     lor = creditscalc.Client('Lorraine', 'Tkachyk', 27)
    #     lor.city = 'Calgary'
    #
    #     self.assertEqual(27, lor.age)
    #     self.assertEqual(29, lar.age)
    #     self.assertEqual('Lorraine Tkachyk', lor.name)
    #     self.assertEqual('Larry Shumlich', lar.name)
    #
    #     lar.first_name = 'Lars'
    #     self.assertEqual('Lars', lar.first_name)
    #     self.assertEqual('Lars Shumlich', lar.name)
    #
    #     lar.last_name = 'Shumy'
    #     self.assertEqual('Shumy', lar.last_name)
    #     self.assertEqual('Lars Shumy', lar.name)
    #
    #     lar.birthday()
    #     self.assertEqual(30, lar.age)
    #
    #     lar.birthday()
    #     lar.birthday()
    #     lar.birthday()
    #     self.assertEqual(33, lar.age)
    #     self.assertEqual(27, lor.age)

    @staticmethod
    def create_test_wb():
        wb = Workbook()
        wb.create_sheet("Rates")
        wb.create_sheet("Facilities")
        wb.create_sheet("Clients")
        wb.create_sheet("2018-07")

        ws = wb["Rates"]

        ws.cell(row=1, column=1).value = 'Location'
        ws.cell(row=2, column=1).value = 'King Edward'
        ws.cell(row=3, column=1).value = 'King Edward'
        ws.cell(row=4, column=1).value = 'King Edward'
        ws.cell(row=5, column=1).value = 'King Edward'
        ws.cell(row=6, column=1).value = 'King Edward'
        ws.cell(row=1, column=2).value = 'Type'
        ws.cell(row=2, column=2).value = 'Desk'
        ws.cell(row=3, column=2).value = 'Gallery'
        ws.cell(row=4, column=2).value = 'Meeting'
        ws.cell(row=5, column=2).value = 'Theatre'
        ws.cell(row=6, column=2).value = 'Outdoor'
        ws.cell(row=1, column=3).value = 'Credits'
        ws.cell(row=2, column=3).value = '1'
        ws.cell(row=3, column=3).value = '5'
        ws.cell(row=4, column=3).value = '3'
        ws.cell(row=5, column=3).value = '20'
        ws.cell(row=6, column=3).value = '20'

        ws = wb["Clients"]

        ws.cell(row=1, column=1).value = 'Name'
        ws.cell(row=2, column=1).value = 'Carri Cordon'
        ws.cell(row=3, column=1).value = 'Caren Alsop'

        ws = wb["Facilities"]

        ws.cell(row=1, column=1).value = 'Room Name'
        ws.cell(row=2, column=1).value = 'First Floor Hall'
        ws.cell(row=3, column=1).value = 'Second Floor Hall'
        ws.cell(row=4, column=1).value = 'Third Floor Hall'
        ws.cell(row=5, column=1).value = 'RGO Treehouse (South)'
        ws.cell(row=6, column=1).value = 'RGO Treehouse (North)'
        ws.cell(row=7, column=1).value = 'RGO Treehouse (Full)'
        ws.cell(row=8, column=1).value = 'Studio (North)'
        ws.cell(row=9, column=1).value = 'Studio (South)'
        ws.cell(row=10, column=1).value = 'ArtPark'
        ws.cell(row=11, column=1).value = 'Desk 3rd floor 1'
        ws.cell(row=12, column=1).value = 'Desk 3rd floor 2'
        ws.cell(row=13, column=1).value = 'Desk 3rd floor 3'
        ws.cell(row=14, column=1).value = 'Desk 3rd floor 4'
        ws.cell(row=15, column=1).value = 'Desk 3rd floor 5'
        ws.cell(row=16, column=1).value = 'Desk 3rd floor 6'
        ws.cell(row=17, column=1).value = 'Desk 3rd floor 7'
        ws.cell(row=18, column=1).value = 'Desk 3rd floor 8'
        ws.cell(row=19, column=1).value = 'Desk 3rd floor 9'
        ws.cell(row=1, column=2).value = 'Type'
        ws.cell(row=2, column=2).value = 'Gallery'
        ws.cell(row=3, column=2).value = 'Gallery'
        ws.cell(row=4, column=2).value = 'Gallery'
        ws.cell(row=5, column=2).value = 'Meeting'
        ws.cell(row=6, column=2).value = 'Meeting'
        ws.cell(row=7, column=2).value = 'Meeting'
        ws.cell(row=8, column=2).value = 'Theatre'
        ws.cell(row=9, column=2).value = 'Theatre'
        ws.cell(row=10, column=2).value = 'Outdoor'
        ws.cell(row=11, column=2).value = 'Desk'
        ws.cell(row=12, column=2).value = 'Desk'
        ws.cell(row=13, column=2).value = 'Desk'
        ws.cell(row=14, column=2).value = 'Desk'
        ws.cell(row=15, column=2).value = 'Desk'
        ws.cell(row=16, column=2).value = 'Desk'
        ws.cell(row=17, column=2).value = 'Desk'
        ws.cell(row=18, column=2).value = 'Desk'
        ws.cell(row=19, column=2).value = 'Desk'
        ws.cell(row=1, column=3).value = 'Location'
        ws.cell(row=2, column=3).value = 'King Edward'
        ws.cell(row=3, column=3).value = 'King Edward'
        ws.cell(row=4, column=3).value = 'King Edward'
        ws.cell(row=5, column=3).value = 'King Edward'
        ws.cell(row=6, column=3).value = 'King Edward'
        ws.cell(row=7, column=3).value = 'King Edward'
        ws.cell(row=8, column=3).value = 'King Edward'
        ws.cell(row=9, column=3).value = 'King Edward'
        ws.cell(row=10, column=3).value = 'King Edward'
        ws.cell(row=11, column=3).value = 'King Edward'
        ws.cell(row=12, column=3).value = 'King Edward'
        ws.cell(row=13, column=3).value = 'King Edward'
        ws.cell(row=14, column=3).value = 'King Edward'
        ws.cell(row=15, column=3).value = 'King Edward'
        ws.cell(row=16, column=3).value = 'King Edward'
        ws.cell(row=17, column=3).value = 'King Edward'
        ws.cell(row=18, column=3).value = 'King Edward'
        ws.cell(row=19, column=3).value = 'King Edward'

        ws = wb["2018-07"]
        ws.cell(row=1, column=1).value = None
        ws.cell(row=2, column=1).value = '2018-07-01'
        ws.cell(row=3, column=1).value = '2018-07-02'
        ws.cell(row=4, column=1).value = '2018-07-03'
        ws.cell(row=5, column=1).value = '2018-07-04'
        ws.cell(row=1, column=2).value = 'Room X'
        ws.cell(row=2, column=2).value = None
        ws.cell(row=3, column=2).value = ''
        ws.cell(row=4, column=2).value = 'Eric'
        ws.cell(row=5, column=2).value = None
        ws.cell(row=1, column=3).value = 'Room Y'
        ws.cell(row=2, column=3).value = None
        ws.cell(row=3, column=3).value = None
        ws.cell(row=4, column=3).value = None
        ws.cell(row=5, column=3).value = None
        ws.cell(row=1, column=4).value = 'Room Z'
        ws.cell(row=2, column=4).value = ''
        ws.cell(row=3, column=4).value = ''
        ws.cell(row=4, column=4).value = ''
        ws.cell(row=5, column=4).value = ''
        ws.cell(row=1, column=5).value = 'Room A'
        ws.cell(row=2, column=5).value = 'Peter'
        ws.cell(row=3, column=5).value = ''
        ws.cell(row=4, column=5).value = ''
        ws.cell(row=5, column=5).value = 'Mule'

        return wb
