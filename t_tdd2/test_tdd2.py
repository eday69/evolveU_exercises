
import unittest
import datetime
import tdd2
from openpyxl import Workbook


class TestTddExcel(unittest.TestCase):

    def test_add_five(self):
        self.assertEqual(8, tdd2.add_five(3))
        self.assertNotEqual(7, tdd2.add_five(5))

    def test_my_max(self):
        self.assertEqual(5, tdd2.my_max([1, 2, 3, 4, 5]))
        self.assertNotEqual(1, tdd2.my_max([1, 2, 3, 4, 5]))

    def test_my_min(self):
        self.assertEqual(1, tdd2.my_min([1, 2, 3, 4, 5]))
        self.assertNotEqual(5, tdd2.my_min([1, 2, 3, 4, 5]))

    def test_has_string(self):
        self.assertEqual(["Mary had"],
                         tdd2.has_string(
                             ["Mary had",
                              "a little lamb",
                              "little lamb",
                              "Whose fleece",
                              ], "Mary"))
        self.assertEqual(["a little lamb", "little lamb"],
                         tdd2.has_string(
                             ["Mary had",
                              "a little lamb",
                              "little lamb",
                              "Whose fleece",
                              ], "lamb"))

    def test_to_date(self):
        dt = tdd2.to_date("2010-08-02")
        self.assertIsInstance(dt, datetime.date)
        self.assertEqual(2010, dt.year)
        self.assertEqual(8, dt.month)
        self.assertEqual(2, dt.day)

    def test_date_diff(self):
        self.assertEqual(1, tdd2.date_diff("2018-09-02", "2018-09-01"))

    def test_days_to_doom(self):
        self.assertEqual(100, tdd2.days_to_doom('2018-09-01'))

    def test_contains(self):
        self.assertTrue(tdd2.contains(['a', 'b', 'd'], "a"))

    def test_add_contents(self):
        self.assertEqual(6, tdd2.add_contents([1, 2, 3]))

    def test_lookup(self):
        self.assertEqual('one mine', tdd2.lookup({1: 'one', 2: 'two', 3: 'three'}, 1))

    def test_cell_not_empty(self):
        self.assertTrue(tdd2.cell_not_empty('Larry'))
        self.assertFalse(tdd2.cell_not_empty(''))
        self.assertFalse(tdd2.cell_not_empty(None))

    def test_check_client(self):
        wb = self.create_test_wb()
        ws = wb["Clients"]

        self.assertTrue(tdd2.check_client(ws, 'Luis'))
        self.assertTrue(tdd2.check_client(ws, 'Eric'))

    def test_verify_clients(self):
        wb = self.create_test_wb()
        self.assertTrue(tdd2.verify_clients(wb, '2018-07'))

    def test_process(self):
        tdd2.process(['moduleName'])
        tdd2.process(['moduleName', '2018-07-01'])
        tdd2.process(['moduleName', '2018-07-04', 'test.xlsx'])

    @staticmethod
    def create_test_wb():
        wb = Workbook()
        wb.create_sheet("Clients")
        wb.create_sheet("2018-07")

        print(wb.sheetnames)

        ws = wb["Clients"]

        ws.cell(row=1, column=1).value = 'Luis'
        ws.cell(row=2, column=1).value = 'Carlos'
        ws.cell(row=3, column=1).value = 'Eric'
        ws.cell(row=4, column=1).value = 'Peter'
        ws.cell(row=5, column=1).value = 'Mule'
        ws = wb["2018-07"]
        ws.cell(row=1, column=1).value = None
        ws.cell(row=2, column=1).value = '2018-07-01'
        ws.cell(row=3, column=1).value = '2018-07-02'
        ws.cell(row=4, column=1).value = '2018-07-03'
        ws.cell(row=5, column=1).value = '2018-07-04'
        ws.cell(row=1, column=2).value = 'Room X'
        ws.cell(row=2, column=2).value = None
        ws.cell(row=3, column=2).value = ''
        ws.cell(row=4, column=2).value = 'Eric'
        ws.cell(row=5, column=2).value = None
        ws.cell(row=1, column=3).value = 'Room Y'
        ws.cell(row=2, column=3).value = None
        ws.cell(row=3, column=3).value = None
        ws.cell(row=4, column=3).value = None
        ws.cell(row=5, column=3).value = None
        ws.cell(row=1, column=4).value = 'Room Z'
        ws.cell(row=2, column=4).value = ''
        ws.cell(row=3, column=4).value = ''
        ws.cell(row=4, column=4).value = ''
        ws.cell(row=5, column=4).value = ''
        ws.cell(row=1, column=5).value = 'Room A'
        ws.cell(row=2, column=5).value = 'Peter'
        ws.cell(row=3, column=5).value = ''
        ws.cell(row=4, column=5).value = ''
        ws.cell(row=5, column=5).value = 'Mule'

        return wb
