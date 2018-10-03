# tdd2.py by Eric Day
import sys
import datetime
import openpyxl


# Load client DB
# Open month tab
# For each row, for each column
#     check client in client list



def add_five(x):
    return x + 5


def my_max(test_list):
    mymax = -sys.maxsize - 1
    for i in test_list:
        if i > mymax:
            mymax = i
    return mymax


def my_min(test_list):
    mymin = sys.maxsize
    for i in test_list:
        if i < mymin:
            mymin = i
    return mymin


def has_string(test_list, search_string):
    result = []
    for i in test_list:
        if search_string in i:
            result.append(i)
    return result


def to_date(the_date):
    date_format = "%Y-%m-%d"
    return datetime.datetime.strptime(the_date, date_format)


def date_diff(date1, date2):
    diff = (to_date(date1) - to_date(date2))
    return diff.days


def days_to_doom(startdate):
    return date_diff('2018-12-10', startdate)


def contains(search_list, search_item):
    return search_item in search_list


def add_contents(content):
    result = 0
    for i in content:
        result += i
    return result


def lookup(search_dictionary, key):
    return search_dictionary[key] + ' mine'


def cell_not_empty(cell):
    if cell is not None and cell.strip():
        return True
    else:
        return False
    # return (cell and cell.strip())


def check_client(ws, client):
    # ws = wb['Clients']
    for row in ws.iter_cols(min_col=1, max_col=1):
        for cell in row:
            if client in cell.value:
                return True
    return False


def verify_clients(wb, tab_date):
    clients = True
    if tab_date in wb.sheetnames:
        ws = wb[tab_date]
        for row in ws.iter_cols(min_col=3, min_row=2):
            for cell in row:
                if cell_not_empty(cell.value):
                    if not check_client(wb['Clients'], cell.value):
                        clients = False
        return clients
    else:
        return None


def open_my_workbook(file):
    wb = openpyxl.load_workbook(file)
    return wb

def process(args):
    # print('From process ---> The Args are:', args)
    # print('Hello world from', __name__, __file__)

    if len(args) < 2:
        print('Please pass a date in the format of yyyy-mm')
        return

    date = args[1]
    if len(args) > 2:
        file_name = args[2]
    else:
        file_name = 'cSpace_Booking.xlsx'

    wb = open_my_workbook(file_name)
    print('All clients are ',verify_clients(wb, date),' for date in', file_name)


if __name__ == '__main__':
    process(sys.argv)
