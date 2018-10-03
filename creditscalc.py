# This has been modified by me after uploading to github
# This will calculate the consumption by client

# creditscalc.py by Eric Day
import openpyxl
import sys


class Room:

    def __init__(self, name, type):
        self.name = name
        self.type = type


class Client:

    def __init__(self, name):
        self.name = name
        self.first_name, self.last_name = name.split(' ')
        self.credits = 0

    def add_credit(self, qty):
        self.credits += qty

    def get_credits(self):
        return self.credits


def load_rooms(ws):
    rooms= {}
    for row in ws.iter_rows(min_col=1, max_col=2, min_row=2):
        # print('Room', row[0].value, row[1].value)
        nr = Room(row[0].value, row[1].value)
        rooms[nr.name] = nr
    return rooms


def get_rate(type, rates):
    return rates[type]


def load_clients(ws):
    clients= {}
    for row in ws.iter_cols(min_col=1, max_col=1, min_row=2):
        for cell in row:
            nc = Client(cell.value)
            # clients[nc.first_name] = nc.__dict__
            clients[nc.first_name] = nc
    return clients


def load_rates(ws):
    tmprate = {}
    for row in ws.iter_rows(min_row=2):
        tmprate[row[1].value] = row[2].value
    return tmprate


def cell_not_empty(cell):
    if cell is not None and cell.strip():
        return True
    else:
        return False
    # return (cell and cell.strip())


def valid_client(clients, needle):
    for key in clients.keys():
        if key == needle:
            return True
    return False


def check_client_use(wb, tab_date, clients, rates, rooms):
    if tab_date in wb.sheetnames:
        ws = wb[tab_date]
        for row in ws.iter_cols(min_col=3):
            room = row[0].value
            if room:
                for cell in row[2:]:
                    if cell_not_empty(cell.value):
                        if valid_client(clients, cell.value):
                            # print(cell.value, room)
                            credit =  rates[rooms[room].type]
                            clients[cell.value].add_credit(credit)
    return None


def print_clients(clients):
    for key, client in clients.items():
        print(client.name, client.get_credits())


# clients[cell.value].add_credit(rates[room])


def open_my_workbook(file):
    wb = openpyxl.load_workbook(file)
    return wb


def process(args):
    if len(args) < 2:
        print('Please pass a date in the format of yyyy-mm')
        return

    date = args[1]
    if len(args) > 2:
        file_name = args[2]
    else:
        file_name = 'cSpace_Booking.xlsx'

    wb = open_my_workbook(file_name)
    ws = wb['Rates']
    rates = load_rates(ws)
    # print(rates)
    ws = wb["Facilities"]
    rooms = load_rooms(ws)
    ws = wb["Clients"]
    clients = load_clients(ws)
    check_client_use(wb, date, clients, rates, rooms)
    print_clients(clients)


if __name__ == '__main__':
    process(sys.argv)
